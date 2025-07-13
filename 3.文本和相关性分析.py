'''
@Project ：PycharmProjects
@File    ：词频分析.py
@IDE     ：PyCharm
@Date    ：2023/5/30 14:34

文本和相关性分析的逻辑
1. 待抽取的词分为N个数组(比如2个 - 供应链、风险）
2. 对年报进行分词，成为一个大数组
3. 按照N个数组内的每个词，寻找每个词出现的位置，构建N个位置数组
4. 对比遍历N个数组，寻找在K个词是否存在其他关联词

'''
import os
import re
import jieba
# import difflib
import openpyxl
import re



# 输入年份区间
start_year = "2009"
end_year = "2024"
# 相关性计算，多少个词内
steps = [5, 10, 15]
# 文件存储路径
work_path="/Users/bl/git/pdftToText/reports"


def clean_excel_text(text):
    """
    移除Excel不允许的控制字符（ASCII 0-31中除\t\n\r外的所有字符）
    """
    if not isinstance(text, str):
        return text
    
    # 正则表达式：匹配所有非法控制字符
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    return cleaned

# 在指定步长里找是否匹配
def check_key_in_else_target(idx, step, keywordsgroup_index):
    result = 0

    for i, target in enumerate(keywordsgroup_index):
        if i != 0:
            found = False
            # 在n步之内找正向index
            for x in range(step):
                next_idx = idx + x + 1
                # 如果找到，设置为true，不继续循环
                try:
                    if target.index(next_idx):
                        found = True
                        break
                except Exception as e:
                    found = False
            # 正向没找到，去负向继续找
            if found == False:
                for x in range(step):
                    next_idx = idx - x - 1
                    # 如果找到，设置为true，不继续循环
                    try:
                        if target.index(next_idx):
                            found = True
                            break
                    except Exception as e:
                        found = False
            # 如果一个idx在第一个目标查找词内就没找到，就不用继续遍历了
            if found == False:
                break
            # 如果找到了就继续找下一个
            # 如果找到最后一个匹配也是能找到，就设置数
            if (i == len(keywordsgroup_index) - 1 & found == True):
                result = 1

    return result

# 计算匹配值
def count_relative(keywordsgroup_index, steps):
    # 统计总相关字出现字数
    related_words_counts = [0] * len(steps)
    # 目标遍历数组，其余数组对其进行比较
    target = keywordsgroup_index[0]
    # 遍历
    for idx in target:
        # 根据不同step进行查找
        for i, step in enumerate(steps):
            check_step_count = check_key_in_else_target(idx, step, keywordsgroup_index)
            related_words_counts[i] = related_words_counts[i] + check_step_count

    return  related_words_counts

# 抽取关键字
def extract_keywords(keywordsGroup, root, file_origin_name):
    filename = os.path.join(root, file_origin_name)

    # 每个分类单独计数
    keywordsgroup_counts = [0] * len(keywordsGroup)
    # 记录查找到的keywords的index
    keywordsgroup_index = [[] for i in range(len(keywordsGroup))]
    # 记录查找到的keywords的key
    keywordsgroup_item = [set() for i in range(len(keywordsGroup))]

    # 统计总字数
    total_words = 0
    # 抽取后的文件的存放路径
    tempPath= os.path.join(root, f"temp")
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            content = f.read()

        # 将多个分组的关键词，全部添加到自定义词典
        for keywords in keywordsGroup:
            for keyword in keywords:
                jieba.add_word(keyword)

        # 使用jieba库进行分词
        words = jieba.cut(content)
        words = [word for word in words if word.strip()]
        # 统计总字数
        total_words = len(words)


        #保存下分词后的数据，主要用于过程检测
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Alice"
        worksheet.append(["分词结果"])
        # 1. 生产一个key与index的字典
        words_dict = {}
        try:
            for idx, item in enumerate(words):
                clean_item = clean_excel_text(item)
                if clean_item == "":
                    continue
                # 写入excel
                worksheet.append([clean_item])
                if clean_item not in words_dict:
                    words_dict[clean_item]=[]
                words_dict[clean_item].append(idx)
        except Exception as e:
            print("生产字典错误", str(e))

        try:
            os.makedirs(tempPath, exist_ok=True)
            temp_file_path = f"{tempPath}/{file_origin_name}分词结果.xlsx"
            workbook.save(temp_file_path)
        except Exception as e:
            print(f"创建文件路径错误: {temp_file_path}")

        # 2. 进行遍历匹配
        for j, keywords in enumerate(keywordsGroup):
            for keyword in keywords:
                # 关键词创建正则表达式，与分词的每个词进行匹配
                # 拥有解决“负债率”和“企业负债率”不一致的问题
                # 如果不考虑多词组的关联计算，其实可以在分词时使用search分词，可以直接拆细词
                # 但是考虑关联，需要计算词的距离，所以无法search分词
                patten = rf"{keyword}"
                for word in words_dict:
                    if(re.search(patten, word)):
                        match_item = words_dict[word]
                        # 找到，该分类的统计数据+1
                        keywordsgroup_counts[j] = keywordsgroup_counts[j] + len(match_item)
                        # 记录该分类的index
                        keywordsgroup_index[j].extend(match_item)
                        # 记录匹配的keyword
                        keywordsgroup_item[j].add(keyword)
        # 进行关联计算
        # 将第一个关键词组依次找出顺序，并于剩余关键词组内找寻相关顺序step步长内是否都有关联
        related_words_counts = count_relative(keywordsgroup_index, steps)

    except FileNotFoundError:
        print(f"文件不存在: {filename}")
    except PermissionError:
        print(f"没有访问权限: {filename}")
    except Exception as e:
        print(f"从文件中获取关键词失败: {filename}", e)
        print(str(e))

    return related_words_counts, keywordsgroup_counts, total_words, keywordsgroup_item


def count_txt_files(folder_path, start_year=None, end_year=None):
    """
    统计指定文件夹及其子文件夹中符合年份要求的所有txt文件数量。
    """
    total_files = 0
    processed_files = 0

    try:
        # 遍历文件夹中的所有文件和子文件夹
        for root, dirs, files in os.walk(folder_path):
            # 根据文件夹路径获取年份信息
            match = re.match(r'.*([12]\d{3}).*', os.path.basename(root))
            if match:
                year = match.group(1)
                if (start_year is not None and int(year) < int(start_year)) or (end_year is not None and int(year) > int(end_year)):
                    # 如果年份不符合要求，则移除该文件夹，防止遍历其中的文件
                    dirs[:] = []
                    continue

            # 遍历当前文件夹中的所有txt文件
            for filename in files:
                if filename.endswith('.txt'):
                    total_files += 1
    except FileNotFoundError:
        print(f"文件夹不存在: {folder_path}")
    except PermissionError:
        print(f"没有访问权限: {folder_path}")
    except Exception as e:
        print(f"统计文件数量失败: {folder_path}")
        print(str(e))

    return total_files

def process_files(folder_path, keywordsGroup, start_year=None, end_year=None):
    """
    处理指定文件夹及其子文件夹中的所有txt文件，提取关键词并统计词频和总字数，将结果存储到Excel表格中。
    """
    try:
        # 创建Excel工作簿和工作表，使用 openpyxl
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "alice"
        row = 1  # openpyxl 行索引从 1 开始
        # 添加Excel表头
        headers = ['股票代码', '公司简称', '年份', '总字数']
        for i, keyword in enumerate(keywordsGroup):
            headers.append(f'关键词类别-{i}的次数')
        for i, step in enumerate(steps):
            headers.append(f'关联词step为{step}的次数')
        for i, keywords in enumerate(keywordsGroup):
            headers.append(f'匹配的关键词-{i}')
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=row, column=col_num, value=header)
        row += 1

        total_files = count_txt_files(folder_path, start_year, end_year)
        processed_files = 0

        try:
            # 遍历文件夹中的所有文件和子文件夹
            for root, dirs, files in os.walk(folder_path):
                # 根据文件夹路径获取年份信息
                match = re.match(r'.*([12]\d{3}).*', os.path.basename(root))
                if match:
                    year = match.group(1)
                    if (start_year is not None and int(year) < int(start_year)) or (end_year is not None and int(year) > int(end_year)):
                        # 如果年份不符合要求，则移除该文件夹，防止遍历其中的文件
                        dirs[:] = []
                        continue

                # 遍历当前文件夹中的所有txt文件
                for filename in files:
                    if filename.endswith('.txt'):
                        # 解析文件名，提取股票代码、公司简称和年份
                        match = re.match(r'^(\d{6})_(.*?)_(\d{4})\.txt$', filename)
                        if match:
                            stock_code = match.group(1)
                            company_name = match.group(2)

                            # 更新进度
                            progress = (processed_files / total_files) * 100
                            if progress == 0:
                                print(f"开始处理{filename}")
                            else:
                                print(f"\r进度: {progress:.2f}%，正在处理{filename}, 已处理{processed_files}个文件，总共{total_files}个文件", end='', flush=True)
                            processed_files += 1

                            # 提取关键词并统计词频和总字数
                            related_words_counts, keywordsgroup_counts, total_words, keywordsgroup_item = extract_keywords(keywordsGroup, root, filename)

                            # 将结果写入Excel表格
                            data = [stock_code, company_name, year, total_words]
                            data.extend(keywordsgroup_counts)
                            data.extend(related_words_counts)
                            for keywords in keywordsgroup_item:
                                data.append(','.join(keywords))
                            for col_num, value in enumerate(data, 1):
                                worksheet.cell(row=row, column=col_num, value=value)
                            row += 1

                            # 每处理指定数目个数据就保存一次Excel文件
                            if processed_files % size == 0:
                                workbook.save(name)
        except FileNotFoundError:
            print(f"文件夹不存在: {folder_path}")
        except PermissionError:
            print(f"没有访问权限: {folder_path}")
        except Exception as e:
            print(f"处理文件失败: {folder_path}")
            print(str(e))

        # 保存Excel文件
        try:
            workbook.save(name)
            print("\nExcel文件保存成功！")
        except FileNotFoundError:
            print(f"保存Excel文件失败: 文件夹不存在")
        except PermissionError:
            print(f"保存Excel文件失败: 没有访问权限")
        except Exception as e:
            print("\n保存Excel文件失败。")
            print(str(e))
    except FileNotFoundError:
        print(f"文件夹不存在: {folder_path}")
    except PermissionError:
        print(f"没有访问权限: {folder_path}")
    except Exception as e:
        print("处理文件失败！")
        print(str(e))


if __name__ == '__main__':
    # 设置要提取的关键词列表
    keywordsGroup = [
        [
            "钻井",
            "能源", "气候", "石油",  "油气",  "燃油",   "汽油",  "柴油", "燃料", 
            "WTI",  "原油", "节能", "蓄能",  "储能", 
            "OPEC",  "水力压裂", 
            "碳", "光伏", "环境",
            "太阳能", "乙醇", "内燃机", "页岩油",  "测井",
            "电动汽车", "天然气",  "温室气体",
            "风力", "风能","自然资源", "化石", "页岩","石化",
            "COGIS", "环保", "油井", "碳氢化合物", "烃", "残油", "渣油",
            "管道", "油管", "电力","电能","电价", "核能", "电气", "发电", "光热", "风电", 
            "水力", "水能","势能" , "核电",  "水电",  "电网", "氢能",  "绿氢",
            "全球变暖",  "COGCC", "混合动力",  "生物质能", 
            "可再生", "压裂开采", 
            "油价", "勘探井",  "地热能",  "海洋能",
            "储量", "温室效应", "京都议定书",
            "堆肥", "煤炭", "煤层", "煤油"
        ],
        [
             "波动", "突变", "不可控", "不可预测", "不可预料", "不可预期", "动荡", "振荡", "复杂", "改变", "多变", "变数",
            "风云变幻", "风云突变", "矛盾突出", "变动", "震荡", "风险", "未知", "无常", "不确定", "不稳",
            "很难", "难以", "无法", "徘徊",
            "负面", "困境", "困难", "困扰", "流失", "垄断", "难题", "失衡", "挑战", "突发", "危机", "考验", "冲击",
            "低迷", "薄弱", "不利", "不足", "恶化", "压力", "严峻", "隐患", "争端", "中断", "矛盾", "打击", "紧张",
            "不平衡", "缺少", "短缺", "危险", "混乱",
            "上升", "上涨", "提高", "增加", "下降", "下滑", "减弱", "不旺", "不振", "相对不足",
            "持续", "出现", "大幅", "小幅",
            "成本", "费用", "需求"
        ]  
    ]
    
    root_folder = f"{work_path}"

    # 输入处理结果的文件名
    name = "新词频分析结果.xlsx"
    # 暂存数目大小，默认为100，尽量别改太小，否则IO压力很大。
    size = 100
    # 处理文件夹中的所有txt文件，并将结果存储到Excel表格中
    try:
        if start_year > end_year:
            print("起始年份不能大于中止年份！！！！！")
        else:
            process_files(root_folder, keywordsGroup, start_year, end_year)
    except Exception as e:
        print("文件处理失败！！")
        print(str(e))

